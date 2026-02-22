import uuid
import tempfile
import os

from app.ingest.chunker import chunk_text
from app.rag.vectorstore import VectorStore


async def ingest_excel(
    content: bytes,
    filename: str,
    vector_store: VectorStore,
    user_id: str = "default"
) -> tuple[str, int]:
    """
    Extract text from Excel/CSV files and ingest into vector store.

    Args:
        content: File bytes
        filename: Original filename
        vector_store: VectorStore instance
        user_id: User ID for data isolation

    Returns:
        tuple: (source_id, number_of_chunks)
    """
    source_id = str(uuid.uuid4())
    all_chunks = []

    # Determine file type
    ext = os.path.splitext(filename)[1].lower()

    if ext == ".csv":
        import csv
        import io
        text_content = content.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text_content))
        rows = list(reader)

        if not rows:
            raise ValueError("CSV file is empty")

        # Convert CSV to readable text
        headers = rows[0] if rows else []
        text_parts = []
        text_parts.append(f"Data from: {filename}\n")
        text_parts.append(f"Columns: {', '.join(headers)}\n")
        text_parts.append(f"Total rows: {len(rows) - 1}\n\n")

        for i, row in enumerate(rows[1:], 1):
            row_text = " | ".join(
                f"{headers[j] if j < len(headers) else f'Col{j+1}'}: {cell}"
                for j, cell in enumerate(row)
                if cell.strip()
            )
            if row_text:
                text_parts.append(f"Row {i}: {row_text}")

        full_text = "\n".join(text_parts)

    elif ext in (".xlsx", ".xls"):
        try:
            import openpyxl
        except ImportError:
            raise ValueError(
                "openpyxl is required for Excel support. Install with: pip install openpyxl"
            )

        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        try:
            wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
            text_parts = [f"Data from: {filename}\n"]

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))

                if not rows:
                    continue

                text_parts.append(f"\n--- Sheet: {sheet_name} ---\n")

                # First row as headers
                headers = [str(cell) if cell is not None else "" for cell in rows[0]]
                text_parts.append(f"Columns: {', '.join(h for h in headers if h)}")
                text_parts.append(f"Total rows: {len(rows) - 1}\n")

                for i, row in enumerate(rows[1:], 1):
                    row_text = " | ".join(
                        f"{headers[j] if j < len(headers) else f'Col{j+1}'}: {str(cell)}"
                        for j, cell in enumerate(row)
                        if cell is not None and str(cell).strip()
                    )
                    if row_text:
                        text_parts.append(f"Row {i}: {row_text}")

            wb.close()
            full_text = "\n".join(text_parts)
        finally:
            os.unlink(tmp_path)
    else:
        raise ValueError(f"Unsupported file type: {ext}. Supported: .csv, .xlsx, .xls")

    if not full_text.strip():
        raise ValueError("No data found in the file")

    # Chunk the text
    chunks = chunk_text(full_text)

    for chunk in chunks:
        chunk["metadata"]["source_id"] = source_id
        chunk["metadata"]["source_name"] = filename
        chunk["metadata"]["source_type"] = "spreadsheet"

    all_chunks.extend(chunks)

    if not all_chunks:
        raise ValueError("No content could be extracted from the file")

    # Add to vector store
    await vector_store.add_documents(all_chunks, source_id, filename, "spreadsheet", user_id)

    return source_id, len(all_chunks)
